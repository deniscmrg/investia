# core/views.py
from datetime import date
import io
import json
import re
from decimal import Decimal, InvalidOperation
import math

from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Q
from django.http import JsonResponse

import yfinance as yf
import pandas as pd
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.core.management import call_command
from django.db import transaction
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from datetime import date

import sys
import os
import requests

from django.conf import settings

from .models import (
    Cliente,
    OperacaoCarteira,
    Acao,
    ImportacaoJob,  # log das importações (mantido)
    Patrimonio,  # <<< tipado
    Custodia,
    Cotacao,
    RecomendacaoDiariaAtual,
    RecomendacaoDiariaAtualNova,
    RecomendacaoIA,
    MT5Order,
    MT5Deal,
    OperacaoMT5Leg,
)
from .serializers import (
    UserSerializer,
    ClienteSerializer,
    OperacaoCarteiraSerializer,
    AcaoSerializer,
    RecomendacaoIASerializer,
)
from .mt5_client import MT5Client, MT5Response
from uuid import uuid4
from datetime import datetime, timedelta
from django.utils.dateparse import parse_datetime


MT5_SUCCESS_RETCODES = {10008, 10009, 10010}
MT5_TP_RETRY_RETCODES = {10016, 10032}
MT5_RETCODE_MESSAGES = {
    10004: "Ordem rejeitada pelo servidor",
    10006: "Negociação desabilitada para o símbolo",
    10016: "Stops inválidos para a ordem",
    10030: "Volume inválido para a ordem",
    10031: "Preço inválido para a ordem",
    10032: "Stop loss ou take profit inválido",
    10033: "Volume insuficiente para execução",
    10034: "Mercado fechado para o símbolo",
}

BCB_SERIES = {
    "ipca": {
        "series": 13522,
        "label": "IPCA (12 meses)",
        "description": "IPCA acumulado em 12 meses (IBGE)",
    },
    "igpm": {
        "series": 189,
        "label": "IGP-M (12 meses)",
        "description": "IGP-M acumulado em 12 meses (FGV)",
    },
    "selic": {
        "series": 432,
        "label": "SELIC",
        "description": "Taxa Selic Meta (BCB)",
    },
}


def _finite_float(value):
    """Converte para float retornando None quando não for número finito."""
    if value is None:
        return None
    try:
        val = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(val) or math.isinf(val):
        return None
    return val


def _fetch_bcb_series_latest(series_id: int) -> tuple[float | None, str | None]:
    """
    Busca o último valor disponível de uma série do Banco Central (SGS).
    Retorna (valor_float, data_str) ou (None, None) em caso de falha.
    """
    url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{series_id}/dados/ultimos/1?formato=json"
    try:
        resp = requests.get(url, timeout=5)
        if not resp.ok:
            return None, None
        data = resp.json()
        if not isinstance(data, list) or not data:
            return None, None
        item = data[0]
        raw_val = item.get("valor")
        raw_date = item.get("data")
        if raw_val is None:
            return None, raw_date
        val = float(str(raw_val).replace(",", "."))
        return val, raw_date
    except (requests.RequestException, ValueError, TypeError):
        return None, None


def _mt5_order_success(retcode: int | None, order_ticket: int | None) -> bool:
    """
    Considera sucesso quando o retcode está entre os códigos de aceite da corretora.
    Caso o retcode venha vazio, garante que ao menos há um order_ticket válido.
    """
    if retcode is None:
        return bool(order_ticket)
    return retcode in MT5_SUCCESS_RETCODES


def _mt5_error_detail(
    *,
    retcode: int | None,
    response_json,
    fallback_error: str | None,
    http_status: int | None = None,
) -> str:
    """
    Gera uma mensagem humana consolidando retcode + informações retornadas pelo MT5.
    """
    pieces: list[str] = []
    if retcode and retcode in MT5_RETCODE_MESSAGES:
        pieces.append(MT5_RETCODE_MESSAGES[retcode])

    if isinstance(response_json, dict):
        for key in ("error", "detail", "message", "comment", "retcode_description"):
            val = response_json.get(key)
            if val:
                pieces.append(str(val))
        # Alguns gateways retornam lista de erros em `errors`
        errors = response_json.get("errors")
        if isinstance(errors, (list, tuple)):
            pieces.extend(str(e) for e in errors if e)
    elif isinstance(response_json, str) and response_json.strip():
        pieces.append(response_json.strip())

    if fallback_error:
        pieces.append(str(fallback_error))

    parts = [p for p in (piece.strip() for piece in pieces) if p]
    if not parts:
        base = "Falha ao enviar ordem"
    else:
        # remove duplicados mantendo ordem
        seen = set()
        dedup = []
        for part in parts:
            if part not in seen:
                dedup.append(part)
                seen.add(part)
        base = "; ".join(dedup)

    suffix = ""
    if retcode:
        suffix = f" (retcode={retcode})"
    elif http_status and http_status >= 400:
        suffix = f" (HTTP {http_status})"

    return f"{base}{suffix}"


def _mt5_should_retry_without_tp(retcode: int | None, detail: str | None) -> bool:
    """
    Determina se vale a pena reenviar ordem sem TP por possível problema de stops.
    """
    if retcode in MT5_TP_RETRY_RETCODES:
        return True
    if detail:
        lowered = detail.lower()
        if "stop" in lowered or "tp" in lowered or "take profit" in lowered:
            return True
    return False


def _coerce_int(value) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _has_open_operacao(cliente: Cliente, base_ticker: str) -> bool:
    try:
        return OperacaoCarteira.objects.filter(
            cliente=cliente, acao__ticker=base_ticker, data_venda__isnull=True
        ).exists()
    except Exception:
        return False


def _position_conflict_response(cliente: Cliente, base_ticker: str):
    """
    Verifica se já existe posição da carteira aberta para o cliente+ticker.
    Retorna Response(409) quando existir conflito.
    """
    if not base_ticker:
        return None
    if _has_open_operacao(cliente, base_ticker):
        msg = f"Cliente já possui posição aberta para {base_ticker}"
        return Response({"detail": msg}, status=status.HTTP_409_CONFLICT)
    return None


def _parse_mt5_order_response(mt5_response: MT5Response | None):
    """
    Padroniza leitura de retcode/order_ticket e prepara detalhes para log/retorno.
    """
    if mt5_response is None:
        return False, None, None, None, "Sem resposta do servidor MT5"

    response_json = (
        mt5_response.data
        if mt5_response.ok
        else (mt5_response.data if mt5_response.data is not None else mt5_response.error)
    )

    order_ticket = None
    retcode = None
    if isinstance(response_json, dict):
        order_ticket = _coerce_int(response_json.get("order") or response_json.get("order_ticket"))
        retcode = _coerce_int(
            response_json.get("retcode")
            or response_json.get("retcode_external")
            or response_json.get("code")
        )
        if retcode == 0:
            retcode = None
    else:
        # quando a API retorna somente o número do ticket
        order_ticket = _coerce_int(response_json)

    success = bool(mt5_response.ok and _mt5_order_success(retcode, order_ticket))
    detail = None
    if not success:
        detail = _mt5_error_detail(
            retcode=retcode,
            response_json=response_json,
            fallback_error=mt5_response.error,
            http_status=getattr(mt5_response, "status", None),
        )

    return success, order_ticket, retcode, response_json, detail



# -------------------
# Autenticação / Perfil
# -------------------

class MyTokenObtainPairView(TokenObtainPairView):
    """Retorna access/refresh e inclui username na resposta."""
    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        if response.status_code == 200:
            data = dict(response.data)
            try:
                user = User.objects.get(username=request.data.get("username"))
                data["username"] = user.username
            except User.DoesNotExist:
                pass
            return Response(data)
        return Response(response.data, status=response.status_code)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def perfil_usuario(request):
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def logout_view(request):
    try:
        refresh_token = request.data["refresh"]
        token = RefreshToken(refresh_token)
        token.blacklist()
        return Response({"detail": "Logout realizado com sucesso."})
    except Exception as e:
        return Response({"error": str(e)}, status=400)


@api_view(["POST"])
def login_view(request):
    username = request.data.get("username")
    password = request.data.get("password")
    user = authenticate(username=username, password=password)
    if user:
        refresh = RefreshToken.for_user(user)
        return Response(
            {
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "username": user.username,
            }
        )
    return Response({"detail": "Invalid credentials"}, status=401)


# -------------------
# Core
# -------------------

class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all().order_by("-criado_em")
    serializer_class = ClienteSerializer
    permission_classes = [permissions.IsAuthenticated]


class OperacaoCarteiraViewSet(viewsets.ModelViewSet):
    """
    ViewSet das operações da carteira.
    Filtra por ?cliente=<id> quando o parâmetro é enviado.
    """
    serializer_class = OperacaoCarteiraSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["cliente"]

    def get_queryset(self):
        qs = OperacaoCarteira.objects.all()
        cliente_id = self.request.query_params.get("cliente")
        if cliente_id:
            try:
                cliente_id_int = int(cliente_id)
            except ValueError:
                return OperacaoCarteira.objects.none()
            qs = qs.filter(cliente=cliente_id_int)
        return qs

    def _reconcile_pending_partial(self, qs):
        try:
            abertos = list(qs.filter(data_venda__isnull=True))
        except Exception:
            return
        if not abertos:
            return
        now = timezone.now()
        lookback = timedelta(minutes=15)
        for op in abertos:
            try:
                req_id = f"op:{op.id}"
                ords = list(MT5Order.objects.filter(cliente=op.cliente, request_id=req_id))
                if not ords:
                    continue
                # já executadas, não precisa reconciliar
                if all(str(getattr(o, "status", "")).lower() == "executada" for o in ords):
                    continue
                # janela temporal para evitar chamadas excessivas
                min_created = None
                for o in ords:
                    dt = getattr(o, "created_at", None)
                    if dt and (min_created is None or dt < min_created):
                        min_created = dt
                if min_created and (now - min_created) > lookback:
                    # se há ordens antigas, ainda assim tenta uma vez reconciliar
                    pass

                ip = _get_cliente_ip(op.cliente)
                if not ip:
                    continue
                mt5 = MT5Client(ip)
                inicio_epoch = int(((min_created or now) - timedelta(hours=1)).timestamp())
                fim_epoch = int((now + timedelta(minutes=5)).timestamp())
                deals_resp = mt5.historico_deals(inicio=inicio_epoch, fim=fim_epoch)
                if not deals_resp.ok or not isinstance(deals_resp.data, list):
                    continue
                deals_by_order = {}
                for d in deals_resp.data:
                    try:
                        order_id = int(d.get("order")) if isinstance(d, dict) else None
                    except Exception:
                        order_id = None
                    if not order_id:
                        continue
                    deals_by_order.setdefault(order_id, []).append(d)

                for o in ords:
                    try:
                        order_id = int(o.order_ticket or 0)
                    except Exception:
                        order_id = 0
                    order_deals = deals_by_order.get(order_id, [])
                    vol_exec = 0.0
                    for d in order_deals:
                        try:
                            vol_exec += float(d.get("volume", 0) or 0)
                        except Exception:
                            pass
                    try:
                        vol_req = float(o.volume_req or 0)
                    except Exception:
                        vol_req = 0.0
                    new_status = None
                    if vol_req > 0 and vol_exec >= vol_req:
                        new_status = "executada"
                    elif vol_exec > 0:
                        new_status = "parcial"
                    else:
                        # mantém pendente/enviada
                        if str(getattr(o, "status", "")).lower() in ("enviada", "pendente", "parcial"):
                            new_status = str(o.status)
                    if new_status and new_status != o.status:
                        o.status = new_status
                        o.save(update_fields=["status", "updated_at"])
            except Exception:
                # não interrompe reconciliação em caso de erro de uma operação
                continue

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        # reconcilia status das operações abertas (pendentes/parciais) via MT5
        self._reconcile_pending_partial(qs)
        return super().list(request, *args, **kwargs)


class AcaoViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Acao.objects.all()
    serializer_class = AcaoSerializer
    permission_classes = [IsAuthenticated]


class RecomendacaoIAViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Lista recomendações da IA direcional.

    Filtros via query params:
        - data (YYYY-MM-DD), default = hoje
        - tipo = compra | venda | todos
        - min_prob = probabilidade mínima (float)
    """

    serializer_class = RecomendacaoIASerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = RecomendacaoIA.objects.select_related("acao")

        # Filtro por data
        data_param = self.request.query_params.get("data")
        if data_param:
            try:
                ref_date = datetime.strptime(data_param, "%Y-%m-%d").date()
            except ValueError:
                ref_date = None
        else:
            ref_date = None

        # Se nenhuma data válida foi informada, usa a última data disponível
        if ref_date is None:
            last_date = (
                RecomendacaoIA.objects.order_by("-data")
                .values_list("data", flat=True)
                .first()
            )
            ref_date = last_date or date.today()

        qs = qs.filter(data=ref_date)

        tipo = (self.request.query_params.get("tipo") or "todos").lower()
        min_prob_param = self.request.query_params.get("min_prob")
        min_prob = None
        if min_prob_param is not None:
            try:
                min_prob = float(min_prob_param)
            except (TypeError, ValueError):
                min_prob = None

        if tipo == "compra":
            qs = qs.filter(classe="UP_FIRST")
            if min_prob is not None:
                qs = qs.filter(prob_up__gte=min_prob)
            qs = qs.order_by("-prob_up")
        elif tipo == "venda":
            qs = qs.filter(classe="DOWN_FIRST")
            if min_prob is not None:
                qs = qs.filter(prob_down__gte=min_prob)
            qs = qs.order_by("-prob_down")
        else:
            # todos
            if min_prob is not None:
                qs = qs.filter(
                    Q(classe="UP_FIRST", prob_up__gte=min_prob)
                    | Q(classe="DOWN_FIRST", prob_down__gte=min_prob)
                    | Q(classe="NONE")
                )
            qs = qs.order_by("-data", "acao__ticker")

        return qs

    def list(self, request, *args, **kwargs):
        """
        Lista recomendações IA já com coluna de variação diária (%)
        baseada no ticker da ação.
        """
        response = super().list(request, *args, **kwargs)

        data = response.data
        if not isinstance(data, list):
            return response

        # coleta tickers base (ex.: PETR4) presentes na resposta
        tickers_base = []
        for item in data:
            ticker = (item.get("acao_ticker") or "").strip().upper()
            if ticker:
                tickers_base.append(ticker)

        variacoes = _variacao_dia_por_ticker(tickers_base)

        for item in data:
            ticker = (item.get("acao_ticker") or "").strip().upper()
            var = variacoes.get(ticker)
            item["variacao_dia"] = float(var) if var is not None else None

        return response


# -------------------
# Cotações helpers (MT5 first, yfinance fallback)
# -------------------

def _mt5_cotacao_info(ip: str | None, base_ticker: str) -> dict[str, float]:
    """Retorna informações de cotação via MT5: preço, máxima e mínima."""
    info: dict[str, float] = {}
    if not ip or not base_ticker:
        return info
    try:
        mt5 = MT5Client(ip)
        q = mt5.cotacao(base_ticker)
        if q.ok and isinstance(q.data, dict):
            data = q.data

            def _first_float(*keys):
                for key in keys:
                    val = _finite_float(data.get(key))
                    if val is not None:
                        return val
                return None

            preco = _first_float("last", "ask", "bid", "price")
            maxima = _first_float(
                "high", "max", "maximum", "high_price", "highPrice", "max_price"
            )
            minima = _first_float(
                "low", "min", "minimum", "low_price", "lowPrice", "min_price"
            )

            if preco is not None:
                info["preco"] = preco
            if maxima is not None:
                info["max"] = maxima
            if minima is not None:
                info["min"] = minima
    except Exception:
        info = {}
    return info


def _mt5_preco_atual(ip: str | None, base_ticker: str) -> float | None:
    """Consulta preço atual via MT5 API do cliente. Retorna last ou ask/bid."""
    return _mt5_cotacao_info(ip, base_ticker).get("preco")


def _variacao_dia_por_ticker(bases: list[str]) -> dict[str, float]:
    """
    Calcula % de variação do dia (preço atual vs abertura) por ticker base.

    Estratégia:
      - tenta obter o preço atual via MT5 usando um IP padrão (quando configurado);
      - sempre busca dados do dia via yfinance para obter a abertura e,
        quando necessário, também o preço atual;
      - retorna apenas entradas válidas (com abertura > 0).
    """
    # normaliza tickers base (ex.: "PETR4")
    bases_norm = sorted({(b or "").strip().upper() for b in bases if b})
    if not bases_norm:
        return {}

    variacoes: dict[str, float] = {}
    preços_atuais: dict[str, float] = {}

    # 1) MT5 (opcional, IP padrão para cotações globais)
    mt5_ip = getattr(settings, "MT5_QUOTES_IP", None) or getattr(
        settings, "MT5_DEFAULT_IP", None
    )
    if mt5_ip:
        for base in bases_norm:
            try:
                info = _mt5_cotacao_info(mt5_ip, base)
                preco = _finite_float(info.get("preco"))
                if preco is not None:
                    preços_atuais[base] = preco
            except Exception:
                # silencioso: fallback total em yfinance
                continue

    # 2) yfinance para abertura (e preço atual quando MT5 não trouxe)
    for base in bases_norm:
        ticker_yf = f"{base}.SA"
        preco_atual = preços_atuais.get(base)
        preco_abertura = None

        try:
            acao = yf.Ticker(ticker_yf)
            info = acao.info or {}
            if preco_atual is None:
                preco_atual = _finite_float(info.get("regularMarketPrice"))
            preco_abertura = _finite_float(info.get("regularMarketOpen"))
        except Exception:
            preco_abertura = None

        if preco_atual is None or not preco_abertura or preco_abertura == 0:
            continue

        try:
            variacao = ((preco_atual - preco_abertura) / preco_abertura) * 100.0
        except ZeroDivisionError:
            continue

        variacoes[base] = variacao

    return variacoes


# -------------------
# Cotações (yfinance)
# -------------------

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def cotacoes_atuais(request):
    """
    Body:
    { "tickers": ["ABEV3.SA", "PETR4.SA"] }
    """
    tickers = request.data.get("tickers", [])
    if not tickers:
        return Response({"error": "Nenhum ticker enviado"}, status=400)

    resultado = {}
    for ticker in tickers:
        try:
            acao = yf.Ticker(ticker)
            info = acao.info
            preco_atual = info.get("regularMarketPrice")
            preco_abertura = info.get("regularMarketOpen")
            variacao_pct = None
            try:
                if preco_atual is not None and preco_abertura not in (None, 0):
                    variacao_pct = (
                        (float(preco_atual) - float(preco_abertura))
                        / float(preco_abertura)
                    ) * 100.0
            except Exception:
                variacao_pct = None

            resultado[ticker] = {
                "preco_atual": preco_atual,
                "preco_abertura": preco_abertura,
                "variacao_pct": variacao_pct,
                "moeda": info.get("currency"),
            }
        except Exception as e:
            resultado[ticker] = {"erro": str(e)}

    return Response(resultado)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def indices_economicos(request):
    """
    Retorna os últimos valores de IPCA, IGP-M e taxa Selic (usada como proxy para SEFIP).
    Fonte: Banco Central do Brasil (API SGS).
    """
    payload = {}
    for key, meta in BCB_SERIES.items():
        valor, data_ref = _fetch_bcb_series_latest(meta["series"])
        payload[key] = {
            "label": meta["label"],
            "description": meta["description"],
            "value": valor,
            "date": data_ref,
        }
    return Response(payload)


# -------------------
# Dashboard RV
# -------------------

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard_rv(request):
    hoje = date.today()

    # Todas as operações em aberto
    posicoes = list(
        OperacaoCarteira.objects.filter(data_venda__isnull=True).select_related("acao", "cliente")
    )

    # Prepara mapas auxiliares para status sem N+1
    op_ids = [op.id for op in posicoes if op.id]
    legs_por_op = set(
        OperacaoMT5Leg.objects.filter(operacao_id__in=op_ids).values_list("operacao_id", flat=True)
    ) if op_ids else set()

    req_ids = [f"op:{op_id}" for op_id in op_ids]
    orders = (
        MT5Order.objects.filter(request_id__in=req_ids).only("status", "request_id")
        if req_ids else []
    )
    statuses_por_op: dict[int, list[str]] = {}
    for ord in orders:
        req = (ord.request_id or "").strip()
        if not req.startswith("op:"):
            continue
        try:
            op_id = int(req.split(":", 1)[1])
        except Exception:
            continue
        statuses_por_op.setdefault(op_id, []).append(str(getattr(ord, "status", "")).lower())

    def _status(op):
        try:
            if getattr(op, "data_venda", None):
                return "encerrada"
            if op.id in legs_por_op:
                return "executada"
            statuses = {s for s in statuses_por_op.get(op.id, []) if s}
            if not statuses:
                return "manual"
            if statuses.issubset({"executada"}):
                return "executada"
            if "executada" in statuses or "parcial" in statuses:
                return "parcial"
            if "pendente" in statuses or "enviada" in statuses:
                return "pendente"
            if statuses.issubset({"rejeitada", "cancelada"}):
                return "falha"
        except Exception:
            pass
        return None

    # Tickers base e um IP MT5 associado (se existir) por ticker
    ticker_ips = {}
    for op in posicoes:
        base = (op.acao.ticker or "").strip().upper()
        if not base:
            continue
        if base in ticker_ips:
            continue
        # prefere IP privado quando configurado
        ip_publico = (getattr(op.cliente, "vm_ip", "") or "").strip()
        ip_privado = (getattr(op.cliente, "vm_private_ip", "") or "").strip()
        ip = ip_privado or ip_publico or None
        if ip:
            ticker_ips[base] = ip

    cotacoes: dict[str, dict[str, float]] = {}
    # 1) MT5 por ticker base
    for base, ip in ticker_ips.items():
        info = _mt5_cotacao_info(ip, base)
        if info:
            cotacoes[f"{base}.SA"] = info

    # 2) fallback yfinance para preencher ausências ou máximas/mínimas
    tickers_all = list({(op.acao.ticker or '').strip().upper() + ".SA" for op in posicoes if op.acao and op.acao.ticker})
    faltando = [
        t
        for t in tickers_all
        if t not in cotacoes
        or any(cotacoes[t].get(k) is None for k in ("preco", "max", "min"))
    ]
    if faltando:
        try:
            data = yf.download(
                tickers=faltando,
                period="1d",
                interval="1d",
                progress=False,
                auto_adjust=False,
            )
            if not isinstance(data, pd.DataFrame) or data.empty:
                raise ValueError("Sem dados do yfinance")

            # Seleciona a última linha disponível
            last_row = data.iloc[-1]

            def _get_value(row, column, ticker):
                try:
                    if isinstance(row.index, pd.MultiIndex):
                        raw = row[(column, ticker)]
                    else:
                        raw = row[column]
                except Exception:
                    return None
                return _finite_float(raw)

            for ticker in faltando:
                entry = cotacoes.setdefault(ticker, {})
                close_val = _get_value(last_row, "Close", ticker)
                high_val = _get_value(last_row, "High", ticker)
                low_val = _get_value(last_row, "Low", ticker)

                if entry.get("preco") is None and close_val is not None:
                    entry["preco"] = close_val
                if entry.get("max") is None and high_val is not None:
                    entry["max"] = high_val
                if entry.get("min") is None and low_val is not None:
                    entry["min"] = low_val
        except Exception:
            pass

    posicionadas = []
    for op in posicoes:
        ticker = (op.acao.ticker or '').strip().upper() + ".SA"
        info = cotacoes.get(ticker, {})
        preco_atual = _finite_float(info.get("preco"))
        preco_max = _finite_float(info.get("max"))
        preco_min = _finite_float(info.get("min"))
        dias_pos = (hoje - op.data_compra).days

        lucro = (
            ((preco_atual - float(op.preco_unitario)) / float(op.preco_unitario)) * 100
            if preco_atual is not None and op.preco_unitario not in (None, 0)
            else None
        )

        posicionadas.append({
            "id": op.id,
            "cliente": op.cliente.nome,
            "cliente_id": op.cliente_id,
            "acao": op.acao.ticker,
            "data_compra": str(op.data_compra),
            "preco_compra": float(op.preco_unitario),
            "quantidade": op.quantidade,
            "valor_total_compra": float(op.valor_total_compra),
            "preco_atual": float(preco_atual) if preco_atual is not None else None,
            "preco_max": float(preco_max) if preco_max is not None else None,
            "preco_min": float(preco_min) if preco_min is not None else None,
            "lucro_percentual": lucro,
            "valor_alvo": float(op.valor_alvo) if op.valor_alvo else None,
            "dias_posicionado": dias_pos,
            "status": _status(op),
        })

    return Response({"posicionadas": posicionadas})

# -------------------
# Importação: Patrimônio / Custódia (tipado)
# -------------------

# normalização de números "pt-BR" -> float
NUM_BR = re.compile(r'^\s*-?\d{1,3}(\.\d{3})*(,\d+)?\s*$')

def _to_number(val):
    if val in (None, ''):
        return None
    if isinstance(val, (int, float, Decimal)):
        return float(val)
    s = str(val).strip()
    if s == '':
        return None
    if NUM_BR.match(s):
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return None


def _to_int(val):
    """
    Conversão resiliente para inteiro: usa a conversão numérica padrão e,
    em caso de falha, tenta extrair apenas dígitos.
    """
    num = _to_number(val)
    if num is not None:
        try:
            return int(num)
        except (TypeError, ValueError, OverflowError):
            pass
    digits = re.sub(r'\D', '', str(val or ''))
    return int(digits) if digits else None

def _xlsx_to_rows(file_bytes: bytes):
    """Converte Excel em lista de dicts (cabeçalho = primeira linha não vazia) usando a primeira aba.

    Mantido para compatibilidade. Para múltiplas abas, usar _xlsx_to_sheets.
    """
    try:
        import pandas as pd
        df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
        df = df.dropna(how='all').dropna(axis=1, how='all')
        df.columns = [str(c).strip() for c in df.columns]
        return df.fillna('').to_dict(orient='records')
    except Exception:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        # acha a primeira linha com algum header
        header_idx = next((i for i, r in enumerate(rows) if any(c not in (None, '') for c in r)), 0)
        headers = [str(h).strip() if h is not None else '' for h in rows[header_idx]]
        out = []
        for r in rows[header_idx+1:]:
            if r is None or all(v in (None, '') for v in r):
                continue
            item = {}
            for i, h in enumerate(headers):
                key = h or f'col_{i+1}'
                val = '' if i >= len(r) or r[i] in (None, '') else r[i]
                item[key] = val
            out.append(item)
        return out


def _xlsx_to_sheets(file_bytes: bytes) -> dict:
    """Lê todas as abas do Excel e retorna {nome_aba: [rows como dict]}.

    - Usa pandas.read_excel(sheet_name=None) quando disponível.
    - Fallback com openpyxl para iterar worksheets.
    """
    sheets = {}
    try:
        import pandas as pd
        dfs = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl", sheet_name=None)
        for name, df in dfs.items():
            if df is None:
                continue
            df = df.dropna(how='all').dropna(axis=1, how='all')
            if df.empty:
                continue
            df.columns = [str(c).strip() for c in df.columns]
            sheets[str(name).strip()] = df.fillna('').to_dict(orient='records')
        return sheets
    except Exception:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header_idx = next((i for i, r in enumerate(rows) if any(c not in (None, '') for c in r)), 0)
                headers = [str(h).strip() if h is not None else '' for h in rows[header_idx]]
                out = []
                for r in rows[header_idx + 1:]:
                    if r is None or all(v in (None, '') for v in r):
                        continue
                    item = {}
                    for i, h in enumerate(headers):
                        key = h or f'col_{i+1}'
                        val = '' if i >= len(r) or r[i] in (None, '') else r[i]
                        item[key] = val
                    out.append(item)
                if out:
                    sheets[str(ws.title).strip()] = out
        except Exception:
            pass
        return sheets

def _norm_key(k: str) -> str:
    return str(k or '').strip().lower()

# mapa de cabeçalhos -> campos dos models
HEADER_MAP_PATRIMONIO = {
    'cod. cliente': 'cod_cliente',
    'cód. cliente': 'cod_cliente',
    'cod_cliente': 'cod_cliente',
    'cod cliente': 'cod_cliente',
    'cód cliente': 'cod_cliente',
    'codigo cliente': 'cod_cliente',
    'código cliente': 'cod_cliente',
    'codigo do cliente': 'cod_cliente',
    'código do cliente': 'cod_cliente',
    'codcliente': 'cod_cliente',
    'codigo assessor': 'codigo_assessor',
    'código assessor': 'codigo_assessor',
    'codigo_assessor': 'codigo_assessor',
    'codigo_acessor': 'codigo_assessor',
    'apelido': 'nome',
    'nome': 'nome',
    'patrimônio total': 'patrimonio_total',
    'patrimonio total': 'patrimonio_total',
    'saldo total': 'saldo_total',
    'garantia utilizada': 'garantia_utilizada',
    'garantias disponíveis': 'garantias_disponiveis',
    'garantias disponiveis': 'garantias_disponiveis',
    'd0': 'd0',
    'd1': 'd1',
    'd2': 'd2',
}

HEADER_MAP_CUSTODIA = {
    'cod. cliente': 'cod_cliente',
    'cód. cliente': 'cod_cliente',
    'cod_cliente': 'cod_cliente',
    'cod cliente': 'cod_cliente',
    'cód cliente': 'cod_cliente',
    'codigo cliente': 'cod_cliente',
    'código cliente': 'cod_cliente',
    'codigo do cliente': 'cod_cliente',
    'código do cliente': 'cod_cliente',
    'codcliente': 'cod_cliente',
    'codigo assessor': 'codigo_assessor',
    'código assessor': 'codigo_assessor',
    'codigo_assessor': 'codigo_assessor',
    'codigo_acessor': 'codigo_assessor',
    'nome': 'nome',
    'apelido': 'nome',
    'ativo': 'ativo',
    'ticker': 'ativo',
    'isin': 'isin',
    'tipo ativo': 'tipo_ativo',
    'quantidade': 'quantidade',
    'preço médio': 'preco_medio',
    'preco medio': 'preco_medio',
    'valor total': 'valor_total',
}

def _map_row(row: dict, tipo: str) -> dict:
    out = {}
    mapa = HEADER_MAP_PATRIMONIO if tipo == 'patrimonio' else HEADER_MAP_CUSTODIA
    for k, v in row.items():
        nk = _norm_key(k)
        if nk not in mapa:
            continue
        field = mapa[nk]
        if field in ('nome', 'ativo', 'isin', 'tipo_ativo'):
            out[field] = (None if v == '' else str(v).strip())
        elif field in ('cod_cliente', 'codigo_assessor'):
            out[field] = _to_int(v)
        else:
            out[field] = _to_number(v)
    # elimina linhas totalmente vazias
    if not any(v is not None and v != '' for v in out.values()):
        return {}
    if tipo == 'patrimonio':
        # força o código do assessor quando não vier no arquivo
        out['codigo_assessor'] = out.get('codigo_assessor') or 22397
    return out


def _infer_tipo(sheet_name: str, rows: list[dict]) -> str | None:
    """Tenta inferir o tipo da aba: 'patrimonio' | 'custodia' | None."""
    name = _norm_key(sheet_name)
    if 'patrim' in name:
        return 'patrimonio'
    if 'custod' in name:
        return 'custodia'

    if not rows:
        return None
    # usa os headers da primeira linha para heurística
    headers = set(_norm_key(k) for k in rows[0].keys())
    p_matches = len(headers.intersection(set(HEADER_MAP_PATRIMONIO.keys())))
    c_matches = len(headers.intersection(set(HEADER_MAP_CUSTODIA.keys())))
    if p_matches == 0 and c_matches == 0:
        return None
    return 'patrimonio' if p_matches >= c_matches else 'custodia'


class ImportacaoUploadView(APIView):
    """
    POST multipart/form-data:
        tipo: 'patrimonio' | 'custodia' | 'auto' (auto = lê todas as abas e importa ambas)
        data_referencia: 'YYYY-MM-DD'
        arquivo: <xlsx>
        force: 'true' (opcional) → apaga a data e recarrega
    Regras:
      - Se já existir dados na data e não vier force, retorna 409 (confirmação).
      - Se force=true, apaga aquela data e insere os novos (por tipo detectado).
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        tipo = (request.data.get('tipo') or '').strip().lower() or 'auto'
        data_ref = request.data.get('data_referencia')
        force = str(request.data.get('force', 'false')).lower() == 'true'
        arq = request.FILES.get('arquivo')

        if tipo not in ('patrimonio', 'custodia', 'auto'):
            return Response({'detail': "Parâmetro 'tipo' inválido. Use 'patrimonio', 'custodia' ou 'auto'."}, status=400)
        if not data_ref:
            return Response({'detail': "Parâmetro 'data_referencia' é obrigatório (YYYY-MM-DD)."}, status=400)
        if not arq:
            return Response({'detail': "Arquivo não enviado (campo 'arquivo')."}, status=400)

        file_bytes = arq.read()

        # Modo AUTO: lê todas as abas e separa por tipo automaticamente
        if tipo == 'auto':
            sheets = _xlsx_to_sheets(file_bytes)
            if not sheets:
                return Response({'detail': 'Arquivo sem abas válidas.'}, status=400)

            registros_por_tipo: dict[str, list[dict]] = {'patrimonio': [], 'custodia': []}
            for sheet_name, linhas in sheets.items():
                detected = _infer_tipo(sheet_name, linhas)
                if detected not in ('patrimonio', 'custodia'):
                    continue
                for r in linhas:
                    m = _map_row(r, detected)
                    if m:
                        m['data_referencia'] = data_ref
                        registros_por_tipo[detected].append(m)

            # remove vazios
            registros_por_tipo = {k: v for k, v in registros_por_tipo.items() if v}
            if not registros_por_tipo:
                return Response({'detail': 'Nenhum registro válido encontrado nas abas.'}, status=400)

            # Checagem de existência por tipo
            conflitos = []
            if 'patrimonio' in registros_por_tipo and Patrimonio.objects.filter(data_referencia=data_ref).exists():
                conflitos.append('patrimonio')
            if 'custodia' in registros_por_tipo and Custodia.objects.filter(data_referencia=data_ref).exists():
                conflitos.append('custodia')

            if conflitos and not force:
                return Response(
                    {
                        'detail': f"Já há dados para esta data: {', '.join(conflitos)}. Confirma sobrescrita?",
                        'need_confirm': True,
                        'tipos_existentes': conflitos,
                    },
                    status=409
                )

            resumo = {}
            with transaction.atomic():
                # Patrimônio
                if 'patrimonio' in registros_por_tipo:
                    status_job = 'ok'
                    if force and Patrimonio.objects.filter(data_referencia=data_ref).exists():
                        Patrimonio.objects.filter(data_referencia=data_ref).delete()
                        status_job = 'sobrescrito'
                    objs = [Patrimonio(**m) for m in registros_por_tipo['patrimonio']]
                    Patrimonio.objects.bulk_create(objs, batch_size=1000)
                    ImportacaoJob.objects.create(
                        tipo='patrimonio',
                        data_referencia=data_ref,
                        total_linhas=len(objs),
                        status=status_job
                    )
                    resumo['patrimonio'] = {'linhas': len(objs), 'status': status_job}

                # Custódia
                if 'custodia' in registros_por_tipo:
                    status_job = 'ok'
                    if force and Custodia.objects.filter(data_referencia=data_ref).exists():
                        Custodia.objects.filter(data_referencia=data_ref).delete()
                        status_job = 'sobrescrito'
                    objs = [Custodia(**m) for m in registros_por_tipo['custodia']]
                    Custodia.objects.bulk_create(objs, batch_size=1000)
                    ImportacaoJob.objects.create(
                        tipo='custodia',
                        data_referencia=data_ref,
                        total_linhas=len(objs),
                        status=status_job
                    )
                    resumo['custodia'] = {'linhas': len(objs), 'status': status_job}

            return Response({'ok': True, 'data': data_ref, 'resumo': resumo})

        # Modo tradicional: apenas um tipo explícito, mas agora escolhe a(s) aba(s) correspondente(s)
        sheets = _xlsx_to_sheets(file_bytes)
        registros = []
        if sheets:
            for sheet_name, linhas in sheets.items():
                detected = _infer_tipo(sheet_name, linhas)
                if detected != tipo:
                    continue
                for r in linhas:
                    m = _map_row(r, tipo)
                    if m:
                        m['data_referencia'] = data_ref
                        registros.append(m)
        else:
            # fallback para leitura simples (primeira aba)
            linhas = _xlsx_to_rows(file_bytes)
            for r in linhas:
                m = _map_row(r, tipo)
                if m:
                    m['data_referencia'] = data_ref
                    registros.append(m)

        if not registros:
            return Response({'detail': 'Nenhum registro válido encontrado no arquivo para o tipo informado.'}, status=400)

        Modelo = Patrimonio if tipo == 'patrimonio' else Custodia
        ja_existe = Modelo.objects.filter(data_referencia=data_ref).exists()

        if ja_existe and not force:
            return Response(
                {'detail': 'Já há dados para esta data. Confirma sobrescrita?', 'need_confirm': True},
                status=409
            )

        with transaction.atomic():
            status_job = 'ok'
            if ja_existe and force:
                Modelo.objects.filter(data_referencia=data_ref).delete()
                status_job = 'sobrescrito'

            if tipo == 'patrimonio':
                objs = [Patrimonio(**m) for m in registros]
                Patrimonio.objects.bulk_create(objs, batch_size=1000)
            else:
                objs = [Custodia(**m) for m in registros]
                Custodia.objects.bulk_create(objs, batch_size=1000)

            ImportacaoJob.objects.create(
                tipo=tipo,
                data_referencia=data_ref,
                total_linhas=len(objs),
                status=status_job
            )

        return Response({'ok': True, 'tipo': tipo, 'data': data_ref, 'linhas': len(registros), 'status': status_job})


#---------------------------------------
#  CALCULOS PARA TELA DE RESUMO
#---------------------------------------
from datetime import date
from decimal import Decimal
from rest_framework.decorators import api_view
from rest_framework.response import Response
import yfinance as yf

from .models import Cliente, OperacaoCarteira, Patrimonio

def _to_decimal(v):
    if v is None: 
        return Decimal("0")
    if isinstance(v, Decimal): 
        return v
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0")

@api_view(["GET"])
def carteira_resumo(request, cliente_id):
    # -------- Cliente ----------
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    operacoes = OperacaoCarteira.objects.filter(cliente=cliente)

    # -------- Patrimônio (tabela 'patrimonio') ----------
    cod_cliente = None
    if getattr(cliente, "codigo_xp", None):
        try:
            cod_cliente = int(str(cliente.codigo_xp).strip())
        except Exception:
            cod_cliente = None

    patrimonio_total = Decimal("0")
    saldo_total = Decimal("0")

    if cod_cliente is not None:
        ultimo = (
            Patrimonio.objects
            .filter(cod_cliente=cod_cliente)
            .order_by("-data_referencia", "-criado_em", "-id")
            .first()
        )
        if ultimo:
            patrimonio_total = _to_decimal(ultimo.patrimonio_total)
            saldo_total = _to_decimal(ultimo.saldo_total)

    percentual = _to_decimal(getattr(cliente, "percentual_patrimonio", 0))

    # -------- Posicionadas (MT5 com fallback yfinance) + POSICIONADO (CUSTO - VALOR_ATUAL) ----------
    posicionadas_valor_atual = Decimal("0")
    posicionadas_custo = Decimal("0")
    abertos = operacoes.filter(data_venda__isnull=True)

    # tenta primeiro via MT5 do próprio cliente
    cotacoes: dict[str, float] = {}
    ip_cli = (getattr(cliente, "vm_private_ip", "") or getattr(cliente, "vm_ip", "") or "").strip() or None
    bases = list({(op.acao.ticker or '').strip().upper() for op in abertos if getattr(op.acao, "ticker", None)})
    if ip_cli and bases:
        for base in bases:
            preco = _mt5_preco_atual(ip_cli, base)
            if preco is not None:
                cotacoes[f"{base}.SA"] = float(preco)

    # fallback yfinance
    tickers = [f"{b}.SA" for b in bases]
    faltando = [t for t in tickers if t not in cotacoes]
    if faltando:
        try:
            data = yf.download(tickers=faltando, period="1d", interval="1m", progress=False)
            if len(faltando) == 1:
                ultimo_close = data["Close"].dropna().iloc[-1]
                cotacoes[faltando[0]] = float(ultimo_close)
            else:
                for t in faltando:
                    ultimo_close = (data["Close"][t] if hasattr(data["Close"], "__getitem__") else data["Close"]).dropna().iloc[-1]
                    cotacoes[t] = float(ultimo_close)
        except Exception:
            pass

    for op in abertos:
        t = f"{op.acao.ticker}.SA"
        # fallback: se faltar cotação, usa preço de compra (rentab=0 nessa linha)
        preco_atual = _to_decimal(cotacoes.get(t, float(op.preco_unitario or 0)))
        qtd = _to_decimal(op.quantidade)
        pu  = _to_decimal(op.preco_unitario)

        posicionadas_valor_atual += (preco_atual * qtd)
        posicionadas_custo       += (pu * qtd)

    posicionadas_valor_atual = posicionadas_valor_atual.quantize(Decimal("0.01"))
    posicionadas_custo       = posicionadas_custo.quantize(Decimal("0.01"))
    posicionado              = (- posicionadas_custo + posicionadas_valor_atual).quantize(Decimal("0.01"))

    # -------- Valor disponível --------
    # regra atual: (patrimônio × %) − valor a mercado das abertas
    valor_disponivel = (
        (patrimonio_total * (percentual / Decimal("100"))).quantize(Decimal("0.01"))
        - posicionadas_valor_atual
    ).quantize(Decimal("0.01"))

    # -------- Realizadas (lucro/prejuízo acumulado) ----------
    realizadas = Decimal("0")
    for op in operacoes.filter(data_venda__isnull=False):
        total_compra = _to_decimal(op.preco_unitario) * _to_decimal(op.quantidade)
        total_venda  = _to_decimal(op.preco_venda_unitario) * _to_decimal(op.quantidade)
        realizadas += (total_venda - total_compra)
    realizadas = realizadas.quantize(Decimal("0.01"))

    # -------- Dias desde a primeira compra ----------
    primeira = operacoes.order_by("data_compra").first()
    dias_total = (date.today() - primeira.data_compra).days if (primeira and primeira.data_compra) else 0

    # -------- Rentabilidade média mensal (meses fechados) ----------
    fechadas = operacoes.filter(data_venda__isnull=False)
    rentab_mensal = None
    if fechadas.exists():
        partes = []
        for op in fechadas:
            if not (op.data_compra and op.data_venda and op.preco_unitario):
                continue
            meses = (op.data_venda.year - op.data_compra.year) * 12 + (op.data_venda.month - op.data_compra.month)
            meses = max(meses, 1)
            pu = _to_decimal(op.preco_unitario)
            pv = _to_decimal(op.preco_venda_unitario)
            if pu == 0:
                continue
            pct = ((pv - pu) / pu) * Decimal("100")
            partes.append(pct / Decimal(meses))
        if partes:
            rentab_mensal = float((sum(partes) / len(partes)).quantize(Decimal("0.01")))

    resumo = {
        "patrimonio": float(patrimonio_total.quantize(Decimal("0.01"))),
        "percentual_patrimonio": float(percentual),
        "valor_disponivel": float(valor_disponivel),
        # total a mercado das abertas
        "posicionadas": float(posicionadas_valor_atual),
        # PnL aberto pela sua fórmula (custo - valor atual)
        "posicionado": float(posicionado),
        "realizadas": float(realizadas),
        "dias_total": int(dias_total),
        "rentabilidade_mensal": rentab_mensal,
        # opcional
        "saldo_total": float(saldo_total.quantize(Decimal("0.01"))) if saldo_total is not None else None,
    }
    return Response(resumo)

@api_view(["GET"])
def carteira_detalhe(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    operacoes = OperacaoCarteira.objects.filter(cliente=cliente)

    # ======================
    # 1. Buscar últimas cotações via MT5 (fallback yfinance)
    # ======================
    abertos = operacoes.filter(data_venda__isnull=True)
    bases = list({(op.acao.ticker or '').strip().upper() for op in abertos if op.acao and op.acao.ticker})
    cotacoes: dict[str, float] = {}
    ip_cli = (getattr(cliente, "vm_private_ip", "") or getattr(cliente, "vm_ip", "") or "").strip() or None
    if ip_cli and bases:
        for base in bases:
            preco = _mt5_preco_atual(ip_cli, base)
            if preco is not None:
                cotacoes[f"{base}.SA"] = float(preco)
    # fallback yfinance
    tickers = [f"{b}.SA" for b in bases]
    faltando = [t for t in tickers if t not in cotacoes]
    if faltando:
        try:
            data = yf.download(tickers=faltando, period="1d", interval="1m", progress=False)
            if len(faltando) == 1:
                ultimo = data["Close"].dropna().iloc[-1]
                cotacoes[faltando[0]] = float(ultimo)
            else:
                for t in faltando:
                    ultimo = (data["Close"][t] if hasattr(data["Close"], "__getitem__") else data["Close"]).dropna().iloc[-1]
                    cotacoes[t] = float(ultimo)
        except Exception:
            pass

    # Enriquecer operações com preço atual
    operacoes_data = []
    for op in operacoes:
        op_dict = OperacaoCarteiraSerializer(op).data
        if not op.data_venda:  # só abertas
            ticker = f"{op.acao.ticker}.SA"
            op_dict["preco_atual"] = cotacoes.get(ticker)
        else:
            op_dict["preco_atual"] = None
        operacoes_data.append(op_dict)

    # ======================
    # 2. Calcular Resumo
    # ======================
    patrimonio = sum((op.preco_unitario or 0) * (op.quantidade or 0) for op in operacoes)
    percentual = cliente.percentual_patrimonio or 0
    valor_disponivel = patrimonio * (percentual / 100)

    posicionadas = sum(
        (cotacoes.get(f"{op.acao.ticker}.SA", op.preco_unitario or 0)) * (op.quantidade or 0)
        for op in abertos
    )

    realizadas = sum(
        ((op.preco_venda_unitario or 0) * (op.quantidade or 0)) - ((op.preco_unitario or 0) * (op.quantidade or 0))
        for op in operacoes.filter(data_venda__isnull=False)
    )

    primeira = operacoes.order_by("data_compra").first()
    dias_total = (date.today() - primeira.data_compra).days if primeira and primeira.data_compra else 0

    realizadas_ops = operacoes.filter(data_venda__isnull=False)
    pct_mensal = None
    if realizadas_ops.exists():
        acumulado = []
        for op in realizadas_ops:
            if not op.data_compra or not op.data_venda or not op.preco_unitario:
                continue
            meses = (op.data_venda.year - op.data_compra.year) * 12 + (op.data_venda.month - op.data_compra.month)
            meses = meses if meses > 0 else 1
            pct = ((op.preco_venda_unitario - op.preco_unitario) / op.preco_unitario) * 100
            acumulado.append(pct / meses)
        if acumulado:
            pct_mensal = round(sum(acumulado) / len(acumulado), 2)

    resumo = {
        "patrimonio": float(patrimonio_total.quantize(Decimal("0.01"))),
        "percentual_patrimonio": float(percentual),
        "valor_disponivel": float(valor_disponivel),
        # total a mercado das abertas (como já vinha antes)
        "posicionadas": float(posicionadas_valor_atual),
        # NOVO: rentabilidade das abertas pela sua fórmula (CUSTO - VALOR_ATUAL)
        "posicionado": float(posicionado),
        "realizadas": float(realizadas.quantize(Decimal("0.01"))),
        "dias_total": int(dias_total),
        "rentabilidade_mensal": rentab_mensal,
    }
    return Response(resumo)

    # ======================
    # 3. Resposta única
    # ======================
    return Response({
        "cliente": {
            "id": cliente.id,
            "nome": cliente.nome,
            "percentual_patrimonio": cliente.percentual_patrimonio,
        },
        "resumo": resumo,
        "operacoes": operacoes_data,
    })


#---- ESSE É PRA GERAR A TELA CONSOLIDADA DE PATRIMÔNIO -------
@api_view(["GET"])
def patrimonio_disponivel(request):
    """
    Consolida por cliente os mesmos cálculos do `carteira_resumo`:
      - patrimonio_total  (tabela Patrimonio, último registro do cliente)
      - posicionadas_valor_atual (a mercado, via yfinance)
      - valor_disponivel = (patrimonio_total * percentual/100) - posicionadas_valor_atual
      - total_consolidado = posicionadas_valor_atual + valor_disponivel
    Retorna: codigo, nome, patrimonio, total_consolidado, valor_disponivel
    """
    clientes = Cliente.objects.all()
    saida = []

    for cliente in clientes:
        # -------- Patrimônio (tabela 'patrimonio') ----------
        cod_cliente = None
        if getattr(cliente, "codigo_xp", None):
            try:
                cod_cliente = int(str(cliente.codigo_xp).strip())
            except Exception:
                cod_cliente = None

        patrimonio_total = Decimal("0")
        saldo_total = Decimal("0")
        if cod_cliente is not None:
            ultimo = (
                Patrimonio.objects
                .filter(cod_cliente=cod_cliente)
                .order_by("-data_referencia", "-criado_em", "-id")
                .first()
            )
            if ultimo:
                patrimonio_total = _to_decimal(ultimo.patrimonio_total)
                saldo_total = _to_decimal(ultimo.saldo_total)

        percentual = _to_decimal(getattr(cliente, "percentual_patrimonio", 0))

        # -------- Posicionadas (a mercado via yfinance) ----------
        posicionadas_valor_atual = Decimal("0")
        posicionadas_custo = Decimal("0")

        abertos = OperacaoCarteira.objects.filter(cliente=cliente, data_venda__isnull=True)

        # tickers únicos
        tickers = list({f"{op.acao.ticker}.SA" for op in abertos if getattr(op.acao, "ticker", None)})

        cotacoes = {}
        if tickers:
            try:
                # usar 1m pra aproximar “a mercado” (igual ao carteira_resumo)
                data = yf.download(tickers=tickers, period="1d", interval="1m", progress=False)

                if len(tickers) == 1:
                    # quando é 1 ticker, yfinance devolve Series
                    ultimo_close = data["Close"].dropna().iloc[-1]
                    cotacoes[tickers[0]] = float(ultimo_close)
                else:
                    # quando são vários, vem DataFrame multi-coluna
                    for t in tickers:
                        serie = data["Close"][t] if isinstance(data["Close"], pd.DataFrame) else data["Close"]
                        ultimo_close = serie.dropna().iloc[-1]
                        cotacoes[t] = float(ultimo_close)
            except Exception as e:
                print("Erro yfinance:", e)

        # acumula valores a mercado e custo
        for op in abertos:
            t = f"{op.acao.ticker}.SA"
            # fallback: se faltar cotação, usa preço de compra
            preco_atual = _to_decimal(cotacoes.get(t, float(op.preco_unitario or 0)))
            qtd = _to_decimal(op.quantidade)
            pu  = _to_decimal(op.preco_unitario)

            posicionadas_valor_atual += (preco_atual * qtd)
            posicionadas_custo       += (pu * qtd)

        # arredondamentos iguais ao carteira_resumo
        posicionadas_valor_atual = posicionadas_valor_atual.quantize(Decimal("0.01"))
        posicionadas_custo       = posicionadas_custo.quantize(Decimal("0.01"))
        # posicionado = (- posicionadas_custo + posicionadas_valor_atual)  # não precisamos enviar aqui

        # -------- Valor disponível (mesma regra do carteira_resumo) --------
        valor_disponivel = (
            (patrimonio_total * (percentual / Decimal("100"))).quantize(Decimal("0.01"))
            - posicionadas_valor_atual
        ).quantize(Decimal("0.01"))

        # -------- Total consolidado (ações a mercado + disponível) --------
        total_consolidado = posicionadas_custo  

        saida.append({
            "codigo": cliente.id,
            "nome": cliente.nome,
            "patrimonio": float(patrimonio_total.quantize(Decimal("0.01"))),
            "total_consolidado": float(total_consolidado),
            "valor_disponivel": float(valor_disponivel),
        })

    return Response(saida)


def _build_mt5_status_url(ip: str) -> str:
    scheme = getattr(settings, "MT5_CLIENT_API_SCHEME", "http").rstrip(":/")
    port = getattr(settings, "MT5_CLIENT_API_PORT", "")
    path = getattr(settings, "MT5_CLIENT_API_STATUS_PATH", "/status")

    base = f"{scheme}://{ip}"
    if port:
        base = f"{base}:{port}"

    if not path.startswith("/"):
        path = f"/{path}"

    return f"{base}{path}"


def _evaluate_mt5_status(payload: dict | None) -> tuple[str, str | None]:
    if not payload:
        return "error", "Resposta vazia"

    terminal = payload.get("terminal") or {}
    if not terminal.get("connected"):
        return "offline", "Terminal MT5 desconectado"
    if terminal.get("trade_allowed") is False:
        return "warning", "Trade não permitido"

    return "online", None


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def clientes_mt5_status(request):
    timeout = getattr(settings, "MT5_CLIENT_API_TIMEOUT", 5)
    resultados = []

    try:
        clientes = Cliente.objects.all()
    except Exception as exc:
        return Response(
            {"error": f"Falha ao carregar clientes: {exc}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    for cliente in clientes:
        ip_publico = (cliente.vm_ip or "").strip()
        ip_privado = (cliente.vm_private_ip or "").strip()
        ip_utilizado = ip_privado or ip_publico

        info = {
            "id": cliente.id,
            "nome": cliente.nome,
            "ip": ip_utilizado or None,
            "status": "missing_ip",
            "detail": None,
            "ping": None,
            "checked_at": timezone.now().isoformat(),
        }

        if not ip_utilizado:
            resultados.append(info)
            continue

        url = _build_mt5_status_url(ip_utilizado)
        try:
            resposta = requests.get(url, timeout=timeout)
            if resposta.ok:
                try:
                    payload = resposta.json()
                except ValueError:
                    payload = None

                status_label, detalhe = _evaluate_mt5_status(payload)
                info["status"] = status_label
                info["detail"] = detalhe

                if payload:
                    terminal = payload.get("terminal") or {}
                    info["ping"] = terminal.get("ping")
            else:
                info["status"] = "error"
                info["detail"] = f"HTTP {resposta.status_code}"
        except requests.Timeout:
            info["status"] = "timeout"
            info["detail"] = "Tempo limite excedido"
        except requests.RequestException as exc:
            info["status"] = "error"
            info["detail"] = str(exc)
        except Exception as exc:
            info["status"] = "error"
            info["detail"] = f"Falha inesperada: {exc}"

        resultados.append(info)

    return Response(resultados)


# =============================
# Recomendações disponíveis
# =============================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def recomendacoes_disponiveis(request, cliente_id: int):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    # ações em aberto do cliente
    abertos = OperacaoCarteira.objects.filter(cliente_id=cliente_id, data_venda__isnull=True).values_list("acao_id", flat=True)
    qs = RecomendacaoDiariaAtualNova.objects.exclude(acao_id__in=list(abertos)).order_by("-probabilidade", "-percentual_estimado", "-data")

    tickers_base = [(r.ticker or "").strip().upper() for r in qs if r.ticker]
    tickers_base = list({t for t in tickers_base if t})
    cotacoes, origens = _cotacoes_cliente(cliente, tickers_base)

    dados = []
    for r in qs:
        base_ticker = (r.ticker or "").strip().upper()
        cot_atual = cotacoes.get(base_ticker)
        alvo_5 = float(cot_atual * 1.05) if cot_atual is not None else None
        dados.append(
            {
                "acao_id": r.acao_id,
                "ticker": r.ticker,
                "empresa": r.empresa,
                "preco_compra": float(r.preco_compra or 0),
                "alvo_sugerido": float(r.alvo_sugerido or 0),
                "percentual_estimado": float(r.percentual_estimado or 0),
                "probabilidade": float(r.probabilidade or 0),
                "cotacao_atual": float(cot_atual) if cot_atual is not None else None,
                "cotacao_origem": origens.get(base_ticker),
                "alvo_sugerido_5pct": alvo_5,
            }
        )
    return Response(dados)


# =============================
# Cotação individual (MT5 + fallback)
# =============================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mt5_cotacao_atual(request, cliente_id: int, ticker: str):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    base = (ticker or "").strip().upper()
    if not base:
        return Response({"detail": "Ticker inválido"}, status=400)

    cotacoes, origens = _cotacoes_cliente(cliente, [base])
    cot = cotacoes.get(base)
    origem = origens.get(base)
    alvo = float(cot * 1.05) if cot is not None else None

    return Response(
        {
            "ticker": base,
            "cotacao": float(cot) if cot is not None else None,
            "alvo_5pct": alvo,
            "origem": origem,
        }
    )


# =============================
# Compra MT5 - validação e envio
# =============================

def _get_cliente_ip(cliente: Cliente) -> str | None:
    ip_publico = (cliente.vm_ip or "").strip()
    ip_privado = (cliente.vm_private_ip or "").strip()
    return ip_privado or ip_publico or None


def _cotacoes_cliente(cliente: Cliente, tickers_base: list[str]) -> tuple[dict[str, float], dict[str, str]]:
    """Retorna cotação atual por ticker base usando MT5 com fallback em yfinance."""
    cotacoes: dict[str, float] = {}
    origens: dict[str, str] = {}

    if not tickers_base:
        return cotacoes, origens

    ip = _get_cliente_ip(cliente)
    if ip:
        for base in tickers_base:
            try:
                preco = _mt5_preco_atual(ip, base)
            except Exception:
                preco = None
            if preco is not None:
                cotacoes[base] = float(preco)
                origens[base] = "mt5"

    faltando = [base for base in tickers_base if base not in cotacoes]
    if not faltando:
        return cotacoes, origens

    tickers_sa = [f"{base}.SA" for base in faltando]
    try:
        data = yf.download(tickers=tickers_sa, period="1d", interval="1m", progress=False)
        if isinstance(data, pd.DataFrame) and not data.empty:
            close = data["Close"]
            if isinstance(close, pd.Series):
                serie = close.dropna()
                if not serie.empty:
                    base = faltando[0]
                    cotacoes[base] = float(serie.iloc[-1])
                    origens[base] = "yfinance"
            elif isinstance(close, pd.DataFrame):
                for base in faltando:
                    try:
                        serie = close[f"{base}.SA"].dropna()
                    except Exception:
                        continue
                    if serie.empty:
                        continue
                    cotacoes[base] = float(serie.iloc[-1])
                    origens[base] = "yfinance"
    except Exception:
        pass

    return cotacoes, origens


def _base_from_symbol(symbol: str) -> str:
    return symbol[:-1] if symbol and symbol.endswith("F") else symbol


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mt5_compra_validar(request, cliente_id: int):
    """
    Valida compra e sugere distribuição fracionária quando necessário.

    Body:
    {
      "ticker": "PETR4",              # base
      "modo": "quantidade"|"valor",
      "quantidade": 150,               # quando modo=quantidade
      "valor": 1000.00,                # quando modo=valor
      "execucao": "mercado"|"limite",
      "preco": 36.20,                  # quando execucao=limite
      "tp": 37.50
    }
    """
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    ip = _get_cliente_ip(cliente)
    if not ip:
        return Response({"detail": "Cliente sem IP configurado"}, status=400)

    body = request.data or {}
    ticker_base = str(body.get("ticker") or "").strip().upper()
    modo = (body.get("modo") or "quantidade").lower()
    quantidade = body.get("quantidade")
    valor = body.get("valor")
    execucao = (body.get("execucao") or "mercado").lower()
    preco = body.get("preco")
    tp = body.get("tp")

    if not ticker_base:
        return Response({"detail": "Parâmetro 'ticker' é obrigatório"}, status=400)
    if modo not in ("quantidade", "valor"):
        return Response({"detail": "Modo inválido"}, status=400)
    if execucao not in ("mercado", "limite"):
        return Response({"detail": "Execução inválida"}, status=400)
    if execucao == "limite" and preco is None:
        return Response({"detail": "Preço é obrigatório para ordem a limite"}, status=400)
    if tp is None:
        return Response({"detail": "TP (alvo) é obrigatório"}, status=400)

    conflict = _position_conflict_response(cliente, ticker_base)
    if conflict:
        return conflict

    mt5 = MT5Client(ip)
    base_info = mt5.simbolo(ticker_base)
    frac_ticker = f"{ticker_base}F"
    frac_info = mt5.simbolo(frac_ticker)

    if not base_info.ok:
        return Response({"detail": f"Símbolo base inválido: {ticker_base}"}, status=400)

    # preço de referência para cálculo por valor
    ref_price = None
    if execucao == "mercado":
        q = mt5.cotacao(ticker_base)
        if q.ok and isinstance(q.data, dict):
            ref_price = float(q.data.get("ask") or q.data.get("last") or 0) or None
    else:
        ref_price = float(preco)

    legs = []
    mensagens = []

    def _get_step(info: dict | None) -> tuple[float, float, float] | None:
        if not info or not isinstance(info.data, dict):
            return None
        d = info.data
        return float(d.get("volume_min", 0)), float(d.get("volume_step", 1)), float(d.get("volume_max", 0))

    base_rules = _get_step(base_info)
    frac_rules = _get_step(frac_info) if frac_info.ok else None

    if modo == "quantidade":
        if quantidade is None:
            return Response({"detail": "Informe a quantidade"}, status=400)
        qtd = float(quantidade)
        if base_rules:
            minv, step, _ = base_rules
            qtd_lote = (qtd // step) * step
            rem = qtd - qtd_lote
        else:
            qtd_lote, rem = 0, qtd

        if qtd_lote > 0:
            legs.append({"symbol": ticker_base, "quantidade": int(qtd_lote)})
        if rem > 0 and frac_rules:
            minf, stepf, _ = frac_rules
            # ajusta restante para step do fracionário
            rem_adj = int(rem // stepf) * stepf
            if rem_adj > 0:
                legs.append({"symbol": frac_ticker, "quantidade": int(rem_adj)})
            else:
                mensagens.append("Quantidade remanescente não respeita step do fracionário")

        if not legs:
            return Response({"detail": "Quantidade não respeita os passos de volume"}, status=400)

    else:  # modo == 'valor'
        if valor is None:
            return Response({"detail": "Informe o valor da compra"}, status=400)
        if not ref_price:
            return Response({"detail": "Sem preço de referência para cálculo"}, status=400)
        val = float(valor)
        qtd_lote = 0
        if base_rules:
            minv, step, _ = base_rules
            custo_lote = ref_price * step
            if custo_lote > 0:
                qtd_lote = int(val // custo_lote) * step
                val -= qtd_lote * ref_price
        if qtd_lote > 0:
            legs.append({"symbol": ticker_base, "quantidade": int(qtd_lote)})

        # fracionário com o restante
        if frac_rules and val > 0:
            minf, stepf, _ = frac_rules
            qtd_frac = int(val // (ref_price * stepf)) * stepf
            if qtd_frac > 0:
                legs.append({"symbol": frac_ticker, "quantidade": int(qtd_frac)})

        if not legs:
            return Response({"detail": "Valor insuficiente para os passos de volume"}, status=400)

    # valida cada perna via /validar-ordem
    validacoes = []
    for lg in legs:
        params = {
            "ticker": lg["symbol"],
            "tipo": "compra",
            "quantidade": lg["quantidade"],
            "execucao": execucao,
        }
        if execucao == "limite":
            params["preco"] = preco
        params["tp"] = tp
        v = mt5.validar_ordem(**params)
        validacoes.append({"symbol": lg["symbol"], "ok": v.ok and isinstance(v.data, dict) and v.data.get("ok", False), "motivo": None if not v.ok else (v.data.get("motivo") if isinstance(v.data, dict) else None)})

    return Response({
        "ticker_base": ticker_base,
        "execucao": execucao,
        "preco": preco,
        "tp": tp,
        "legs_sugeridas": legs,
        "validacoes": validacoes,
        "mensagens": mensagens,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mt5_compra(request, cliente_id: int):
    """
    Envia as ordens (1..N) com TP quando possível. Grava MT5Order e retorna group_id + order_tickets.

    Body:
    {
      "ticker_base": "PETR4",
      "execucao": "mercado"|"limite",
      "preco": 36.20,              # quando execucao=limite
      "tp": 37.50,
      "legs": [ {"symbol":"PETR4","quantidade":100}, {"symbol":"PETR4F","quantidade":7} ]
    }
    """
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    ip = _get_cliente_ip(cliente)
    if not ip:
        return Response({"detail": "Cliente sem IP configurado"}, status=400)

    body = request.data or {}
    ticker_base = str(body.get("ticker_base") or "").strip().upper()
    execucao = (body.get("execucao") or "mercado").lower()
    preco = body.get("preco")
    tp = body.get("tp")
    legs = body.get("legs") or []

    if not ticker_base or not legs:
        return Response({"detail": "Parâmetros inválidos"}, status=400)
    if execucao not in ("mercado", "limite"):
        return Response({"detail": "Execução inválida"}, status=400)
    if execucao == "limite" and preco is None:
        return Response({"detail": "Preço é obrigatório para ordem a limite"}, status=400)
    if tp is None:
        return Response({"detail": "TP (alvo) é obrigatório"}, status=400)

    conflict = _position_conflict_response(cliente, ticker_base)
    if conflict:
        return conflict

    mt5 = MT5Client(ip)
    group_id = uuid4()

    # pega login da conta (best-effort)
    acc = mt5.status()
    account_login = None
    try:
        if acc.ok and isinstance(acc.data, dict):
            account_login = (acc.data.get("conta") or {}).get("login")
    except Exception:
        account_login = None

    # cria placeholder em OperacaoCarteira imediatamente (vinculado via request_id nos MT5Order)
    placeholder_op = None
    try:
        acao = Acao.objects.get(ticker=ticker_base)
    except Acao.DoesNotExist:
        acao = None

    total_qty = 0
    try:
        total_qty = sum(int(lg.get("quantidade") or 0) for lg in (legs or []) if isinstance(lg, dict))
    except Exception:
        total_qty = 0

    # preço estimado: limite usa preco informado; mercado tenta ask/last; fallback 0
    ref_price = None
    try:
        if execucao == "limite" and preco is not None:
            ref_price = float(preco)
        else:
            q = mt5.cotacao(ticker_base)
            if q.ok and isinstance(q.data, dict):
                ref_price = float(q.data.get("ask") or q.data.get("last") or 0) or None
    except Exception:
        ref_price = None

    if acao is not None and total_qty > 0:
        try:
            pu = Decimal(str(ref_price if ref_price is not None else 0)).quantize(Decimal("0.01"))
            vtc = (pu * Decimal(str(total_qty))).quantize(Decimal("0.01"))
            placeholder_op = OperacaoCarteira(
                cliente=cliente,
                acao=acao,
                data_compra=timezone.now().date(),
                preco_unitario=pu,
                quantidade=int(total_qty),
                valor_total_compra=vtc,
                valor_alvo=(Decimal(str(tp)).quantize(Decimal("0.01")) if tp is not None else None),
            )
            # tabela é unmanaged; assume existir no banco
            placeholder_op.save(force_insert=True)
        except Exception:
            placeholder_op = None

    results = []
    aggregated_errors = []
    for lg in legs:
        symbol = lg.get("symbol")
        qtd = lg.get("quantidade")
        if not symbol or not qtd:
            continue
        payload = {
            "ticker": symbol,
            "tipo": "compra",
            "quantidade": float(qtd),
            "execucao": execucao,
            "tp": float(tp),
        }
        if execucao == "limite":
            payload["preco"] = float(preco)

        # tenta enviar com TP e aplica fallback sem TP quando houver falha por stops
        apply_tp_after_exec = False
        first_response = mt5.enviar_ordem(payload)
        success, order_ticket, retcode, response_json, error_detail = _parse_mt5_order_response(first_response)

        final_response_json = response_json
        final_retcode = retcode
        final_order_ticket = order_ticket
        final_status = "enviada" if success else "rejeitada"

        # fallback sem TP caso necessário
        if not success and "tp" in payload:
            if (not first_response.ok) or _mt5_should_retry_without_tp(retcode, error_detail):
                alt_payload = dict(payload)
                alt_payload.pop("tp", None)
                apply_tp_after_exec = True
                second_response = mt5.enviar_ordem(alt_payload)
                success2, order_ticket2, retcode2, response_json2, error_detail2 = _parse_mt5_order_response(second_response)
                if success2:
                    success = True
                    final_order_ticket = order_ticket2
                    final_retcode = retcode2
                    final_response_json = response_json2
                    final_status = "enviada"
                    error_detail = None
                else:
                    final_retcode = retcode2 or retcode
                    final_response_json = response_json2
                    error_detail = error_detail2 or error_detail

        order_record = MT5Order.objects.create(
            group_id=group_id,
            cliente=cliente,
            base_ticker=ticker_base,
            symbol=symbol,
            lado="compra",
            execucao=execucao,
            volume_req=qtd,
            price_req=preco if execucao == "limite" else None,
            tp_req=tp,
            apply_tp_after_exec=apply_tp_after_exec,
            account_login=account_login,
            order_ticket=final_order_ticket,
            retcode=final_retcode,
            response_json=str(final_response_json) if final_response_json is not None else None,
            request_id=(f"op:{placeholder_op.id}" if placeholder_op else None),
            status=final_status,
        )

        result_entry = {
            "symbol": symbol,
            "status": final_status,
            "order_ticket": final_order_ticket,
            "retcode": final_retcode,
        }

        if error_detail:
            result_entry["detail"] = error_detail
            aggregated_errors.append(
                {
                    "symbol": symbol,
                    "detail": error_detail,
                    "retcode": final_retcode,
                }
            )

        results.append(result_entry)

    any_success = any(r.get("status") == "enviada" for r in results)
    if not any_success and placeholder_op:
        try:
            MT5Order.objects.filter(group_id=group_id).update(request_id=None)
        except Exception:
            pass
        try:
            placeholder_op.delete()
        except Exception:
            pass
        placeholder_op = None

    response_payload = {
        "group_id": str(group_id),
        "results": results,
        "has_errors": bool(aggregated_errors),
    }
    if aggregated_errors:
        response_payload["errors"] = aggregated_errors

    if aggregated_errors:
        status_code = (
            status.HTTP_400_BAD_REQUEST if len(aggregated_errors) == len(results) else status.HTTP_207_MULTI_STATUS
        )
    else:
        status_code = status.HTTP_200_OK
    return Response(response_payload, status=status_code)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mt5_compra_status(request, cliente_id: int, group_id):
    """
    Verifica status das ordens de um group_id. Quando todas executadas, cria OperacaoCarteira consolidada
    e configura TP via SLTP se necessário.
    """
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    ip = _get_cliente_ip(cliente)
    if not ip:
        return Response({"detail": "Cliente sem IP configurado"}, status=400)

    mt5 = MT5Client(ip)
    orders = list(MT5Order.objects.filter(cliente=cliente, group_id=group_id))
    if not orders:
        return Response({"detail": "group_id desconhecido"}, status=404)

    # verifica deals por ordem
    agora = timezone.now()
    inicio_epoch = int((min(o.created_at for o in orders) - timedelta(hours=1)).timestamp())
    fim_epoch = int((agora + timedelta(minutes=5)).timestamp())

    deals_resp = mt5.historico_deals(inicio=inicio_epoch, fim=fim_epoch)
    deals_by_order: dict[int, list] = {}
    deals_by_symbol: dict[str, list] = {}
    if deals_resp.ok and isinstance(deals_resp.data, list):
        for d in deals_resp.data:
            order_id = None
            symbol_key = None
            if isinstance(d, dict):
                try:
                    order_id = int(d.get("order")) if d.get("order") else None
                except Exception:
                    order_id = None
                symbol_key = d.get("symbol") or None
            if order_id:
                deals_by_order.setdefault(order_id, []).append(d)
            if symbol_key:
                deals_by_symbol.setdefault(symbol_key, []).append(d)

    # posições atuais (fallback quando o histórico ainda não refletiu os negócios)
    posicoes_resp = mt5.posicoes()
    posicoes_por_symbol: dict[str, dict[str, float | list]] = {}
    if posicoes_resp.ok and isinstance(posicoes_resp.data, list):
        for pos in posicoes_resp.data:
            if not isinstance(pos, dict):
                continue
            symbol = (pos.get("symbol") or "").strip()
            if not symbol:
                continue
            volume_val = pos.get("volume")
            if volume_val in (None, ""):
                volume_val = pos.get("volume_float")
            try:
                volume = float(volume_val or 0)
            except Exception:
                volume = 0.0
            price_val = pos.get("price_open") or pos.get("price")
            try:
                price = float(price_val) if price_val is not None else None
            except Exception:
                price = None

            entry = posicoes_por_symbol.setdefault(
                symbol,
                {"volume": 0.0, "prices": []},
            )
            entry["volume"] = float(entry["volume"]) + volume
            if price is not None:
                entry["prices"].append(price)

    # atualiza status local e consolida volumes
    executed_all = True
    summary = []
    for o in orders:
        o_status = o.status
        order_id_key = _coerce_int(o.order_ticket)
        order_deals = deals_by_order.get(order_id_key or 0, [])
        if not order_deals:
            order_deals = deals_by_symbol.get(o.symbol, [])

        vol_exec = 0.0
        vwap_num = 0.0
        for d in order_deals:
            vol = float(d.get("volume", 0))
            price = float(d.get("price", 0))
            vol_exec += vol
            vwap_num += vol * price
            # upsert MT5Deal
            try:
                deal_ticket = int(d.get("ticket"))
            except Exception:
                deal_ticket = None
            if deal_ticket:
                MT5Deal.objects.get_or_create(
                    deal_ticket=deal_ticket,
                    defaults={
                        "cliente": cliente,
                        "order_ticket": int(d.get("order") or 0) or None,
                        "position_ticket": int(d.get("position_id") or d.get("position") or 0) or None,
                        "symbol": d.get("symbol"),
                        "lado": "compra" if int(d.get("type", 0)) in (0, 2) else "venda",  # Buy/Buy limit/stop
                        "volume": vol,
                        "price": price,
                        "commission": float(d.get("commission", 0) or 0),
                        "swap": float(d.get("swap", 0) or 0),
                        "profit": float(d.get("profit", 0) or 0),
                        "time": datetime.fromtimestamp(int(d.get("time")), tz=timezone.utc),
                        "magic": int(d.get("magic", 0) or 0) or None,
                        "comment": d.get("comment"),
                        "raw_json": json.dumps(d),
                    }
                )

        pos_entry = posicoes_por_symbol.get(o.symbol)
        pos_volume = float(pos_entry["volume"]) if pos_entry else 0.0
        pos_avg = None
        if pos_entry and pos_entry["prices"]:
            pos_avg = sum(pos_entry["prices"]) / len(pos_entry["prices"])

        required_volume = float(o.volume_req)
        avg = (vwap_num / vol_exec) if vol_exec > 0 else None

        if vol_exec < required_volume and pos_volume >= required_volume:
            vol_exec = pos_volume
            if pos_avg is not None:
                vwap_num = pos_volume * pos_avg
                avg = pos_avg

        if required_volume > 0 and vol_exec >= required_volume:
            if o_status != "executada":
                o.status = "executada"
                o.save(update_fields=["status", "updated_at"])
            summary.append(
                {
                    "symbol": o.symbol,
                    "executada": True,
                    "volume": vol_exec,
                    "preco_medio": avg,
                    "status": o.status,
                    "retcode": o.retcode,
                }
            )
        else:
            executed_all = False
            if o_status in ("rejeitada", "cancelada"):
                current_status = o_status
            else:
                current_status = "parcial" if vol_exec > 0 else "pendente"
                if o_status != current_status:
                    o.status = current_status
                    o.save(update_fields=["status", "updated_at"])
            summary_entry = {
                "symbol": o.symbol,
                "executada": False,
                "volume_exec": vol_exec,
                "status": current_status,
                "retcode": o.retcode,
            }
            if current_status == "rejeitada" and o.response_json:
                summary_entry["detail"] = o.response_json
            summary.append(summary_entry)

    # quando todas executadas, cria OperacaoCarteira consolidada e aplica TP via SLTP (se necessário)
    created_operacao = None
    if executed_all:
        base = _base_from_symbol(orders[0].symbol)
        # soma volumes e vwap por symbol
        legs_vwap = {}
        for s in summary:
            if not s.get("executada"):
                continue
            symbol = s["symbol"]
            legs_vwap[symbol] = {
                "volume": float(s["volume"]),
                "preco_medio": float(s.get("preco_medio") or 0),
            }

        total_vol = sum(v["volume"] for v in legs_vwap.values())
        if total_vol > 0:
            vwap_total = sum(v["volume"] * v["preco_medio"] for v in legs_vwap.values()) / total_vol
        else:
            vwap_total = 0.0

        # acha acao base
        try:
            acao = Acao.objects.get(ticker=base)
        except Acao.DoesNotExist:
            acao = None

        # tenta localizar placeholder previamente criado via request_id
        placeholder_id = None
        try:
            for o in orders:
                rid = getattr(o, "request_id", None)
                if isinstance(rid, str) and rid.startswith("op:"):
                    try:
                        placeholder_id = int(rid.split(":", 1)[1])
                        break
                    except Exception:
                        placeholder_id = None
        except Exception:
            placeholder_id = None

        if acao is not None:
            op = None
            if placeholder_id:
                try:
                    op = OperacaoCarteira.objects.get(pk=placeholder_id, cliente=cliente)
                except OperacaoCarteira.DoesNotExist:
                    op = None
            if op is None:
                op = OperacaoCarteira(
                    cliente=cliente,
                    acao=acao,
                    data_compra=timezone.now().date(),
                    preco_unitario=Decimal(str(vwap_total)).quantize(Decimal("0.01")),
                    quantidade=int(total_vol),
                    valor_total_compra=Decimal(str(vwap_total * total_vol)).quantize(Decimal("0.01")),
                    valor_alvo=Decimal(str(orders[0].tp_req or 0)).quantize(Decimal("0.01")),
                )
                op.save(force_insert=True)
            else:
                # atualiza placeholder com valores executados
                try:
                    op.preco_unitario = Decimal(str(vwap_total)).quantize(Decimal("0.01"))
                    op.quantidade = int(total_vol)
                    op.valor_total_compra = Decimal(str(vwap_total * total_vol)).quantize(Decimal("0.01"))
                    if getattr(orders[0], "tp_req", None) is not None:
                        op.valor_alvo = Decimal(str(orders[0].tp_req)).quantize(Decimal("0.01"))
                    op.save(update_fields=["preco_unitario", "quantidade", "valor_total_compra", "valor_alvo"])
                except Exception:
                    pass
            created_operacao = op.id

            # vincula legs com position_ticket corrente
            # tenta obter posições para cada symbol
            pos = mt5.posicoes()
            pos_map = {}
            if pos.ok and isinstance(pos.data, list):
                for p in pos.data:
                    if isinstance(p, dict):
                        pos_map[p.get("symbol")] = p
            for sym, data_leg in legs_vwap.items():
                p = pos_map.get(sym)
                position_ticket = int(p.get("ticket")) if p else None
                try:
                    # evita duplicadas se o status for consultado repetidas vezes
                    from django.db.models import Q
                    exists = OperacaoMT5Leg.objects.filter(operacao=op, symbol=sym).exists()
                    if not exists:
                        OperacaoMT5Leg.objects.create(
                            operacao=op,
                            symbol=sym,
                            position_ticket=position_ticket or 0,
                            volume=Decimal(str(data_leg["volume"])),
                            price_open=Decimal(str(data_leg["preco_medio"])),
                            order_ticket=next((int(o.order_ticket or 0) for o in orders if o.symbol == sym), None),
                            deal_tickets=None,
                        )
                except Exception:
                    pass

            # aplica TP via SLTP se alguma ordem marcou apply_tp_after_exec
            if any(o.apply_tp_after_exec for o in orders) and created_operacao:
                alvo = float(orders[0].tp_req or 0)
                for sym in legs_vwap.keys():
                    p = pos_map.get(sym)
                    if p and alvo > 0:
                        mt5.ajustar_stop({"ticket": int(p.get("ticket")), "stop_gain": alvo})

    return Response({
        "executed_all": executed_all,
        "summary": summary,
        "created_operacao_id": created_operacao,
    })


    


# =============================
# Venda MT5 - enviar e monitorar
# =============================

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mt5_venda(request, cliente_id: int, operacao_id: int):
    """
    Envia ordens de venda (mercado/limite) para encerrar 100% da operação indicada (inclui fracionárias).

    Body:
    { "execucao": "mercado"|"limite", "preco": 10.00? }
    """
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    try:
        operacao = OperacaoCarteira.objects.get(pk=operacao_id, cliente=cliente, data_venda__isnull=True)
    except OperacaoCarteira.DoesNotExist:
        return Response({"detail": "Operação não encontrada ou já encerrada"}, status=404)

    ip = _get_cliente_ip(cliente)
    if not ip:
        return Response({"detail": "Cliente sem IP configurado"}, status=400)

    body = request.data or {}
    execucao = (body.get("execucao") or "mercado").lower()
    preco = body.get("preco")
    if execucao not in ("mercado", "limite"):
        return Response({"detail": "Execução inválida"}, status=400)
    if execucao == "limite" and preco is None:
        return Response({"detail": "Preço é obrigatório para ordem a limite"}, status=400)

    mt5 = MT5Client(ip)

    # Buscar legs vinculadas; se inexistentes, usa posições por símbolo base/frac
    base = operacao.acao.ticker
    frac = f"{base}F"

    legs = list(OperacaoMT5Leg.objects.filter(operacao=operacao))

    pos_map = {}
    pos_resp = mt5.posicoes()
    if pos_resp.ok and isinstance(pos_resp.data, list):
        for p in pos_resp.data:
            if isinstance(p, dict) and p.get("symbol"):
                pos_map[p.get("symbol")] = p

    symbols_to_close = []
    if legs:
        # fecha os símbolos presentes nas legs ainda existentes/positivas
        for leg in legs:
            sym = leg.symbol
            p = pos_map.get(sym)
            if p and float(p.get("volume", 0) or 0) > 0:
                symbols_to_close.append((sym, float(p.get("volume"))))
    else:
        # fallback por símbolo base/fracionário
        for sym in (base, frac):
            p = pos_map.get(sym)
            if p and float(p.get("volume", 0) or 0) > 0:
                symbols_to_close.append((sym, float(p.get("volume"))))

    if not symbols_to_close:
        return Response({"detail": "Nenhuma posição MT5 encontrada para encerrar"}, status=400)

    # pega login da conta (best-effort)
    acc = mt5.status()
    account_login = None
    try:
        if acc.ok and isinstance(acc.data, dict):
            account_login = (acc.data.get("conta") or {}).get("login")
    except Exception:
        account_login = None

    group_id = uuid4()
    results = []
    aggregated_errors = []
    for sym, vol in symbols_to_close:
        payload = {
            "ticker": sym,
            "tipo": "venda",
            "quantidade": float(vol),
            "execucao": execucao,
        }
        if execucao == "limite":
            payload["preco"] = float(preco)

        response = mt5.enviar_ordem(payload)
        success, order_ticket, retcode, response_json, error_detail = _parse_mt5_order_response(response)
        status_label = "enviada" if success else "rejeitada"

        MT5Order.objects.create(
            group_id=group_id,
            cliente=cliente,
            base_ticker=base,
            symbol=sym,
            lado="venda",
            execucao=execucao,
            volume_req=vol,
            price_req=preco if execucao == "limite" else None,
            tp_req=None,
            apply_tp_after_exec=False,
            account_login=account_login,
            order_ticket=order_ticket,
            retcode=retcode,
            response_json=str(response_json) if response_json is not None else None,
            status=status_label,
            comment=f"sell_op:{operacao.id}",
        )

        result_entry = {
            "symbol": sym,
            "order_ticket": order_ticket,
            "status": status_label,
            "retcode": retcode,
        }
        if error_detail:
            result_entry["detail"] = error_detail
            aggregated_errors.append({"symbol": sym, "detail": error_detail, "retcode": retcode})

        results.append(result_entry)

    response_payload = {
        "group_id": str(group_id),
        "results": results,
        "has_errors": bool(aggregated_errors),
    }
    if aggregated_errors:
        response_payload["errors"] = aggregated_errors

    if aggregated_errors:
        status_code = (
            status.HTTP_400_BAD_REQUEST if len(aggregated_errors) == len(results) else status.HTTP_207_MULTI_STATUS
        )
    else:
        status_code = status.HTTP_200_OK

    return Response(response_payload, status=status_code)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mt5_venda_status(request, cliente_id: int, group_id):
    """
    Verifica status das ordens de venda de um group_id. Quando todas executadas, atualiza OperacaoCarteira
    (data_venda, preco_venda_unitario, valor_total_venda).
    """
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({"detail": "Cliente não encontrado"}, status=404)

    ip = _get_cliente_ip(cliente)
    if not ip:
        return Response({"detail": "Cliente sem IP configurado"}, status=400)

    orders = list(MT5Order.objects.filter(cliente=cliente, group_id=group_id))
    if not orders:
        return Response({"detail": "group_id desconhecido"}, status=404)

    # tenta inferir operacao_id do comment (sell_op:<id>)
    operacao_id = None
    for o in orders:
        c = (o.comment or "").strip()
        if c.startswith("sell_op:"):
            try:
                operacao_id = int(c.split(":", 1)[1])
                break
            except Exception:
                pass

    try:
        operacao = OperacaoCarteira.objects.get(pk=operacao_id, cliente=cliente)
    except Exception:
        operacao = None

    mt5 = MT5Client(ip)

    # busca deals por período abrangendo as ordens
    agora = timezone.now()
    inicio_epoch = int((min(o.created_at for o in orders) - timedelta(hours=1)).timestamp())
    fim_epoch = int((agora + timedelta(minutes=5)).timestamp())

    deals_resp = mt5.historico_deals(inicio=inicio_epoch, fim=fim_epoch)
    deals_by_order: dict[int, list] = {}
    deals_by_symbol: dict[str, list] = {}
    if deals_resp.ok and isinstance(deals_resp.data, list):
        for d in deals_resp.data:
            order_id = None
            symbol_key = None
            if isinstance(d, dict):
                try:
                    order_id = int(d.get("order")) if d.get("order") else None
                except Exception:
                    order_id = None
                symbol_key = d.get("symbol") or None
            if order_id:
                deals_by_order.setdefault(order_id, []).append(d)
            if symbol_key:
                deals_by_symbol.setdefault(symbol_key, []).append(d)

    executed_all = True
    summary = []
    vwap_num_total = 0.0
    vol_total_exec = 0.0
    last_time = None

    for o in orders:
        o_status = o.status
        order_id_key = _coerce_int(o.order_ticket)
        order_deals = deals_by_order.get(order_id_key or 0, [])
        if not order_deals:
            order_deals = deals_by_symbol.get(o.symbol, [])
        vol_exec = 0.0
        vwap_num = 0.0
        for d in order_deals:
            try:
                vol = float(d.get("volume", 0))
                price = float(d.get("price", 0))
                vol_exec += vol
                vwap_num += vol * price
                ts = int(d.get("time"))
                dt = datetime.fromtimestamp(ts, tz=timezone.utc)
                if (last_time is None) or (dt > last_time):
                    last_time = dt
            except Exception:
                pass

        if vol_exec >= float(o.volume_req):
            if o_status != "executada":
                o.status = "executada"
                o.save(update_fields=["status", "updated_at"])
            avg = (vwap_num / vol_exec) if vol_exec > 0 else None
            summary.append(
                {
                    "symbol": o.symbol,
                    "executada": True,
                    "volume": vol_exec,
                    "preco_medio": avg,
                    "status": o.status,
                    "retcode": o.retcode,
                }
            )
        else:
            executed_all = False
            if o_status in ("rejeitada", "cancelada"):
                current_status = o_status
            else:
                current_status = "pendente"
                if o_status != current_status:
                    o.status = current_status
                    o.save(update_fields=["status", "updated_at"])
            summary_entry = {
                "symbol": o.symbol,
                "executada": False,
                "volume_exec": vol_exec,
                "status": current_status,
                "retcode": o.retcode,
            }
            if current_status == "rejeitada" and o.response_json:
                summary_entry["detail"] = o.response_json
            summary.append(summary_entry)

        vwap_num_total += vwap_num
        vol_total_exec += vol_exec

    updated_operacao = None
    if executed_all and operacao is not None:
        from decimal import Decimal
        vwap_total = (vwap_num_total / vol_total_exec) if vol_total_exec > 0 else 0.0
        venda_data = (last_time.date() if last_time else timezone.now().date())
        try:
            operacao.preco_venda_unitario = Decimal(str(vwap_total)).quantize(Decimal("0.01"))
            operacao.valor_total_venda = (operacao.preco_venda_unitario * Decimal(str(operacao.quantidade))).quantize(Decimal("0.01"))
            operacao.data_venda = venda_data
            operacao.save(update_fields=["preco_venda_unitario", "valor_total_venda", "data_venda"])
            updated_operacao = operacao.id
        except Exception:
            pass

    return Response({
        "executed_all": executed_all,
        "summary": summary,
        "updated_operacao_id": updated_operacao,
    })


def _run_recomendacoes(top_n: int | None = None):
    stream = io.StringIO()
    kwargs = {"stdout": stream, "stderr": stream}
    if top_n is not None:
        kwargs["top"] = top_n
    try:
        call_command("gerar_recomendacoes_intraday", **kwargs)
        return True, stream.getvalue()
    except Exception as exc:
        return False, f"Falha ao executar: {exc}"

@api_view(["GET", "POST"]) 
@permission_classes([IsAuthenticated])
def recomendacoes_api(request):
    """
    GET: retorna dados da vw_recomendacoes_diarias_atual_nova (view "Nova")
    POST: dispara a rotina _run_recomendacoes()
    """
    from .models import RecomendacaoDiariaAtualNova  # usa a view Nova
    if request.method == "POST":
        top_param = request.data.get("top") or request.query_params.get("top")
        try:
            top_value = int(top_param) if top_param is not None else None
        except (TypeError, ValueError):
            top_value = None
        ok, msg = _run_recomendacoes(top_value)
        status = "success" if ok else "error"
        return JsonResponse({"status": status, "message": msg})

    qs = RecomendacaoDiariaAtualNova.objects.all().order_by("-data")

    def _f(x):
        try:
            return float(x)
        except Exception:
            return None

    # calcula variação diária por ticker (atual vs abertura)
    tickers_base = [
        (r.ticker or "").strip().upper()
        for r in qs
        if getattr(r, "ticker", None)
    ]
    variacoes_map = _variacao_dia_por_ticker(tickers_base)

    dados = []
    for r in qs:
        base_ticker = (r.ticker or "").strip().upper()
        variacao_dia = variacoes_map.get(base_ticker)
        dados.append({
            "acao_id": r.acao_id,
            "ticker": r.ticker,
            "empresa": r.empresa,
            "setor": r.setor,
            "data": r.data,
            "preco_compra": _f(r.preco_compra),
            "alvo_sugerido": _f(r.alvo_sugerido),
            "percentual_estimado": _f(r.percentual_estimado),
            "probabilidade": _f(r.probabilidade),
            "vezes_atingiu_alvo_1m": r.vezes_atingiu_alvo_1m,
            "cruza_medias": bool(r.cruza_medias) if r.cruza_medias is not None else None,
            "obv_cres": bool(r.obv_cres) if r.obv_cres is not None else None,
            "vol_acima_media": bool(r.vol_acima_media) if r.vol_acima_media is not None else None,
            # wma602 é decimal nessa view
            "wma602": _f(r.wma602),
            # Campos adicionais da view Nova
            "MIN": _f(getattr(r, "MIN", None)),
            "MAX": _f(getattr(r, "MAX", None)),
            "ALTA": _f(getattr(r, "ALTA", None)),
            "BAIXA": _f(getattr(r, "BAIXA", None)),
            "AMPLITUDE": _f(getattr(r, "AMPLITUDE", None)),
            "AMP_AxF": _f(getattr(r, "AMP_AxF", None)),
            "AMP_MXxMN": _f(getattr(r, "AMP_MXxMN", None)),
            "A_x_F": _f(getattr(r, "A_x_F", None)),
            "ALVO": _f(getattr(r, "ALVO", None)),
            "variacao_dia": _f(variacao_dia) if variacao_dia is not None else None,
        })

    return JsonResponse(dados, safe=False)
